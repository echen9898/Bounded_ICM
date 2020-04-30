
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border, Font
from datetime import datetime
import pytz

# column indexes for all entry keys (0 indexed)
COLS = {
    'date':0,
    'usertag':1,
    'description':2,
    'plots':3,
    'demos':4,
    'experiment_id':5,
    'params_id':6,
    'git_commit':7
}

# row indexes of interest
ROWS = {
    'header':35
}

# dictionary of custom styles
standard = Side(border_style='hair', color='FF000000')
STYLES = {
    'header_alignment':Alignment(vertical='top', horizontal='left'),
    'header_font':Font(name='Calibri', size=11, color='FF000000', bold=True),
    'header_border':Border(left=standard, right=standard, top=standard, bottom=standard, outline=standard),
    'entry_alignment':Alignment(vertical='top', horizontal='left', wrap_text=True),
    'entry_border':Border(left=standard, right=standard, top=standard, bottom=standard, outline=standard),
    'entry_font':Font(name='Calibri', size=11, color='FF000000')
}

def wrap_print(text):
    print('\n')
    print('-'*60)
    print(text)
    print('-'*60)
    print('\n')


def format(sheet):
    ''' Apply default styling to the entire spreadsheet '''
    row_index = 0
    for row in sheet.rows:
        if row_index == ROWS['header']: row_type = 'header'
        else: row_type = 'entry'
        for cell in row:
            cell.font = STYLES['{}_font'.format(row_type)]
            cell.alignment = STYLES['{}_alignment'.format(row_type)]
            cell.border = STYLES['{}_border'.format(row_type)]
        row_index += 1

def row_generator(sheet):
    ''' A generator for rows in an excel sheet '''
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]

def row_matrix(sheet, rows=(None, None)):
    ''' Returns an excel sheet in matrix format. You can
    optionally specify a range (in rows) and only this range
    will be returned
    '''
    table = list()
    for row in sheet.iter_rows(min_row=rows[0], max_row=rows[1]):
        table.append([cell.value for cell in row])
    return table

def get_row_index(query, registry_path, column=None):
    ''' Return the row index (NOT 0 indexed) where a specific query was found. 
    Takes in a string query, a path to the registry file, and the user can
    optionally specify a specific column to search in if this is known
    ahead of time.
    '''
    book = load_workbook(registry_path, read_only=True)
    table = row_generator(book['Experiments'])
    for i, elements in enumerate(table):
        if column is not None and elements[column] == query: return i+1
        elif column is None and query in elements: return i+1
    return None

def get_value(usertag, key, registry_path):
    ''' Returns usertag specific information by key. '''
    book = load_workbook(registry_path, read_only=True)
    table = row_matrix(book['Experiments'])
    row_index = get_row_index(usertag, registry_path, 1)
    if row_index is None: return None
    return table[row_index-1][COLS[key]]

def get_count(query, registry_path, verbatim=True, finetuning=True):
    ''' Returns the number of rows in which this query appears. If verbatim option
    is True, then query 'abc' would yield 0 even if 'abcd' is in the log. If verbatim
    is False, then if any element contains the entry it will be counted. Used
    primarily to determine the seed number on repeated experiments
    '''
    book = load_workbook(registry_path, read_only=True)
    table = row_generator(book['Experiments'])
    count = 0
    for row in table:
        if verbatim:
            if query in row: count += 1
        else:
            row_strings = list()
            # convert from utf-8 to string
            for el in row:
                try: 
                    row_strings.append(el.encode('utf-8'))
                except: 
                    continue
            # check converted strings
            for el in row_strings:
                if not finetuning:
                    if query in el and '_ft' not in el:
                        count += 1
                else:
                    if query in el:
                        count += 1
    return count

def create_usertag(args):
    ''' Create a usertag with the following format: alg_setting_# (e.g. icm_dense_3) '''

    # Rerunning experiment
    if args.tag:
        usertag = args.tag.split('.')
        if '_ft' in usertag: # finetuning
            usertag[-1] = str(get_count(usertag[0]+'.'+usertag[1]+'.', args.registry, verbatim=False))
        else:
            usertag[-1] = str(get_count(usertag[0]+'.', args.registry, verbatim=False, finetuning=False)) # don't forget the decimal point (crucial lol)
        return '.'.join(usertag)

    # Finetuning experiment
    if args.pretrain:
        pretrain_tag = args.pretrain.split('/')[-3].split('_')
        algo = pretrain_tag[0]
        if 'very' in args.env_id.lower(): 
            setting = pretrain_tag[1] + '_ftVerySparse'
        elif 'sparse' in args.env_id.lower(): 
            setting = pretrain_tag[1 ] + '_ftSparse'
        elif 'mario' in args.env_id.lower(): print("NOT IMPLEMENTED CHECK CREATE USERTAG")
        else: print("NOT IMPLEMENTED CHECK CREATE USERTAG")

    # New experiment
    else:
        # Algorithm choice (none, icm, icmpix)
        if args.unsup == None: algo = 'none'
        elif args.unsup == 'action': algo = 'icm'
        elif args.unsup == 'action_lstm': algo = 'lstmpred'
        elif 'state' in args.unsup.lower(): algo = 'icmpix'

        # Reward setting (dense, sparse, verySparse)
        if 'doom' in args.env_id.lower():
            if 'labyrinth' in args.env_id.lower(): setting='labyrinth'
            elif 'very' in args.env_id.lower(): setting='verySparse'
            elif 'sparse' in args.env_id.lower(): setting='sparse'
            else: setting='dense'
        else:
            setting = args.env_id

    # Get trial number using heading count in top rows
    book = load_workbook(args.registry, read_only=True)
    table = row_matrix(book['Experiments'])
    row_index = get_row_index('{}_{}'.format(algo, setting), args.registry)
    trial_number = str(int(table[row_index-1][1]) + 1)

    # Insert the pretrain tags trial/seed number back into the setting
    if args.pretrain:
        setting = setting.split('_')
        setting.insert(1, pretrain_tag[2])
        setting = '_'.join(setting)

    # Unique count id
    if args.pretrain:
        # need the full setting - e.g. icm_labyrinth_20.0 to get the right count
        seed = get_count('{}_{}_{}'.format(algo, setting, trial_number), args.registry, verbatim=False)
    else:
        seed = get_count('{}_{}_{}'.format(algo, setting, trial_number), args.registry)

    usertag = '{}_{}_{}.{}'.format(algo, setting, trial_number, seed)
    return usertag

def dict_to_command(args, store_true_args, default_params, mode):
    ''' Convert a dictionary into a sequence of command line arguments '''
    cmd = ''
    for argument in args:
        value = args[argument]
        if argument not in default_params:
            continue
        if argument in store_true_args:
            if value in {'True', True}:
                cmd += '--{} '.format(argument.replace('_', '-'))
            continue
        cmd += '--{} {} '.format(argument.replace('_', '-'), value)
    return cmd

def update_experiment_count(usertag, registry_path):
    ''' Update the experiment counter in the registry file '''
    usertag_pieces = usertag.split('_')
    trial_number = int(float(usertag_pieces.pop(-1))) # increment the count by one
    alg_and_setting = '_'.join(usertag_pieces)
    row_index = get_row_index(alg_and_setting, registry_path)
    book = load_workbook(registry_path)
    sheet = book['Experiments']
    sheet.cell(row=row_index, column=2).value = trial_number
    book.save(registry_path)

def update_registry(args, usertag, seed_num, exp_id, params_id):
    ''' Update the experiment registry '''

    if args.tag is None and not args.pretrain: # only count new trials (not seeds)
        update_experiment_count(usertag, args.registry)

    book = load_workbook(args.registry)
    sheet = book['Experiments']
    row_index = get_row_index(usertag, args.registry)
    time = datetime.strftime(datetime.now(pytz.timezone('US/Eastern')), "%m-%d-%y %H:%M")
    description = raw_input('ENTER A SHORT EXPERIMENT DESCRIPTION: ')
    branch = subprocess.check_output(['git', 'rev-parse', '--abbrev-ref', 'HEAD']).strip().decode('utf-8')
    commit = subprocess.check_output(['git', 'rev-parse', 'HEAD']).strip().decode('utf-8')
    sheet.append([time, usertag, description, '', '', exp_id, params_id, '{}_{}'.format(branch, commit)])

    format(sheet)
    book.save(args.registry)

    # create a usertag.txt file in tmp with the correct usertag
    with open('./curiosity/src/tmp/usertag.txt', 'w+') as file:
        file.write('{}'.format(usertag))

def numeric(chars):
    ''' Returns true if string is numeric, else false '''
    try:
        float(chars)
        return True
    except ValueError:
        return False

